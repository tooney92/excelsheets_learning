import openpyxl, os

os.chdir('C:/Users/anthony.ime/Downloads')
workbook = openpyxl.load_workbook('Work_Force_Attendance (9) (1).xlsx')
print(workbook.sheetnames)
sheet1 = workbook[workbook.sheetnames[0]]
employees = []
pin = []
department = []
clock_in = []
clock_out = []
total = []
count = 3
for num in range(3, 48):
    employees.append(sheet1.cell(row=num, column=2).value)
    pin.append(int(sheet1.cell(row=3, column=3).value))
    department.append(sheet1.cell(row=num, column=5).value)
    clock_in.append(sheet1.cell(row=num, column=6).value)
    clock_out.append(sheet1.cell(row=num, column=7).value)
    total.append(sheet1.cell(row=num, column=8).value)
    
print(pin)


import openpyxl, os

os.chdir('C:/Users/anthony.ime/Downloads')
workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.sheetnames)
sheet1 = workbook[workbook.sheetnames[0]]
samp = []
for num in range(1, 7):
    samp.append(sheet1.cell(row=num, column=2).value)
print(samp)
